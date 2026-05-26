from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


CORE_VARIABLES = [
    "Lighting",
    "Air compressors",
    "Motor drives",
    "Fans and pumps",
    "Low-enthalpy heat",
]
OTHER_ROW_NAME = "Others (sum mean)"

# Update these sets/prefixes if IDEES introduces new palette variants.
COLOR_GROUPS = {
    "parent": {"4F81BD", "006EBE"},
    "detail": {"974706", "964605", "963732"},
}
COLOR_GROUP_PREFIXES = {
    "parent": ("4F81", "006E"),
    "detail": ("9747", "9646", "9637"),
}


def _default_jrc_idees_path(jrc_idees_path=None):
    if jrc_idees_path is not None:
        return Path(jrc_idees_path)
    base_dir = Path(__file__).resolve().parents[1]
    return base_dir / "Data" / "General" / "JRC-IDEES"


def _clean_str(value):
    if isinstance(value, str):
        return value.strip()
    return None


def _to_float(value):
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _extract_year_columns(df):
    year_columns = []
    years = []
    for c in range(1, df.shape[1]):
        value = df.iat[0, c]
        if pd.isna(value):
            continue
        if isinstance(value, (int, float)):
            year_columns.append(c)
            years.append(int(value))
    return year_columns, years


def _row_values(df, row_idx, col_indexes):
    values = []
    for c in col_indexes:
        value = df.iat[row_idx, c]
        if pd.isna(value):
            values.append(None)
        elif isinstance(value, (int, float)):
            values.append(float(value))
        else:
            values.append(value)
    while values and values[-1] is None:
        values.pop()
    return values


def _font_rgb_6(color):
    if color is None:
        return None
    if color.type != "rgb" or not color.rgb:
        return None
    rgb = str(color.rgb).upper()
    if len(rgb) >= 6:
        return rgb[-6:]
    return None


def _label_color(ws, row_idx):
    if ws is None:
        return None
    # DataFrame row index is 0-based, openpyxl rows are 1-based.
    return _font_rgb_6(ws.cell(row=row_idx + 1, column=1).font.color)


def _color_group(color_hex):
    if not color_hex:
        return None
    clean = str(color_hex).upper()[-6:]
    for group_name, values in COLOR_GROUPS.items():
        if clean in values:
            return group_name
    for group_name, prefixes in COLOR_GROUP_PREFIXES.items():
        if any(clean.startswith(prefix) for prefix in prefixes):
            return group_name
    return None


def _select_non_core_rows_by_color(row_items):
    """
    Color rule for non-core rows:
    - blue + next blue  -> keep blue row
    - blue + next brown -> keep following brown rows until next blue
    - any other case    -> keep current row
    """
    selected = []
    i = 0
    while i < len(row_items):
        current = row_items[i]
        current_group = _color_group(current["color"])

        if current_group == "parent":
            if i + 1 < len(row_items):
                next_group = _color_group(row_items[i + 1]["color"])
                if next_group == "parent":
                    selected.append(current)
                    i += 1
                    continue
                if next_group == "detail":
                    j = i + 1
                    while j < len(row_items) and _color_group(row_items[j]["color"]) == "detail":
                        selected.append(row_items[j])
                        j += 1
                    i = j
                    continue

            selected.append(current)
            i += 1
            continue

        selected.append(current)
        i += 1

    return selected


def _extract_block_variables(df, ws, start_row, end_row, col_indexes):
    core_rows = []
    non_core_rows = []

    for r in range(start_row, end_row + 1):
        label = _clean_str(df.iat[r, 0])
        if not label:
            continue

        item = {
            "label": label,
            "values": _row_values(df, r, col_indexes),
            "color": _label_color(ws, r),
        }
        if label in CORE_VARIABLES:
            core_rows.append(item)
        else:
            non_core_rows.append(item)

    variables = {}
    for item in core_rows:
        variables[item["label"]] = item["values"]

    selected_non_core = _select_non_core_rows_by_color(non_core_rows)
    for item in selected_non_core:
        variables[item["label"]] = item["values"]

    return variables


def _iter_fec_sheet_names(xls):
    index_sheet_name = next(
        (name for name in xls.sheet_names if name.strip().lower() == "index"),
        None,
    )
    if index_sheet_name is None:
        return []

    index_df = pd.read_excel(xls, sheet_name=index_sheet_name, header=None)
    available_sheets = set(xls.sheet_names)
    fec_sheet_names = []
    for value in index_df.to_numpy().ravel():
        clean = _clean_str(value)
        if (
            clean
            and clean.endswith("_fec")
            and clean != "Ind_Summary_fec"
            and clean in available_sheets
            and clean not in fec_sheet_names
        ):
            fec_sheet_names.append(clean)
    return fec_sheet_names


def _find_market_rows(df):
    rows = []
    for r in range(df.shape[0]):
        label = _clean_str(df.iat[r, 0])
        if label and "market shares of energy uses" in label.lower():
            rows.append(r)
    return rows


def _find_next_energy_row(df, start_row):
    for r in range(start_row + 1, df.shape[0]):
        label = _clean_str(df.iat[r, 0])
        if label and label.lower().startswith("energy intensity"):
            return r
    return None


def _previous_non_empty_label_row(df, start_row, min_row):
    for r in range(start_row, min_row - 1, -1):
        label = _clean_str(df.iat[r, 0])
        if label:
            return r
    return None


def extract_subsector_blocks(df, ws=None):
    """
    Extract blocks between 'Market shares...' and 'Energy intensity...'.
    A block starts at each 'Lighting' row and uses the previous non-empty row
    as subsector name.
    """
    col_indexes, years = _extract_year_columns(df)
    if not col_indexes:
        return []

    blocks = []
    for market_row in _find_market_rows(df):
        energy_row = _find_next_energy_row(df, market_row)
        if energy_row is None:
            continue

        lighting_meta = []
        for r in range(market_row + 1, energy_row):
            label = _clean_str(df.iat[r, 0])
            if label == "Lighting":
                subsector_row = _previous_non_empty_label_row(df, r - 1, market_row + 1)
                if subsector_row is None:
                    continue
                lighting_meta.append(
                {
                    "lighting_row": r,
                    "subsector_name_row": subsector_row,
                    "subsector_name": _clean_str(df.iat[subsector_row, 0]),
                    "subsector_total": _row_values(df, subsector_row, col_indexes),
                }
            )

        for idx, info in enumerate(lighting_meta):
            start_row = info["lighting_row"]
            if idx + 1 < len(lighting_meta):
                end_row = lighting_meta[idx + 1]["subsector_name_row"] - 1
            else:
                end_row = energy_row - 1

            if end_row < start_row:
                continue

            variables = _extract_block_variables(df, ws, start_row, end_row, col_indexes)

            if not variables:
                continue

            blocks.append(
                {
                    "years": years,
                    "variables": variables,
                    "subsector_name": info["subsector_name"],
                    "subsector_total": info["subsector_total"],
                }
            )

    return blocks


def read_country_sector_blocks(country_code, jrc_idees_path=None):
    """
    Return extracted subsector blocks grouped by sector (_fec sheet base name).
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    country_dir = jrc_idees_path / country_code
    if not country_dir.exists():
        raise FileNotFoundError(f"No existe la carpeta de pais: {country_dir}")

    industry_files = sorted(country_dir.glob("*Industry*.xlsx"))
    if not industry_files:
        return {}

    industry_file = industry_files[0]
    blocks_by_sector = {}

    with pd.ExcelFile(industry_file) as xls:
        wb = load_workbook(industry_file, data_only=True)
        for fec_sheet_name in _iter_fec_sheet_names(xls):
            df = pd.read_excel(xls, sheet_name=fec_sheet_name, header=None)
            ws = wb[fec_sheet_name] if fec_sheet_name in wb.sheetnames else None
            blocks = extract_subsector_blocks(df, ws=ws)
            sector_name = fec_sheet_name.replace("_fec", "")
            blocks_by_sector[sector_name] = blocks
        wb.close()

    return blocks_by_sector


def aggregate_sector_blocks_to_six_rows_with_total(blocks):
    """
    Build 6 attributes per sector:
    - 5 core variables averaged across subsectors
    - 1 'Others (sum mean)' row:
      sum of non-core variables per subsector, then average across subsectors
    Then add a final 'Total (%)' row as row-wise sum of those 6 attributes.
    """
    if not blocks:
        return None

    years = blocks[0]["years"]
    core_sums = {name: [0.0] * len(years) for name in CORE_VARIABLES}
    other_sums = [0.0] * len(years)
    weight_sums = [0.0] * len(years)

    for block in blocks:
        variables = block["variables"]
        subsector_total = block.get("subsector_total", [])
        for i, _year in enumerate(years):
            weight = _to_float(subsector_total[i] if i < len(subsector_total) else 1.0)
            if weight <= 0.0:
                continue

            weight_sums[i] += weight
            for core_name in CORE_VARIABLES:
                values = variables.get(core_name, [])
                value = values[i] if i < len(values) else 0.0
                core_sums[core_name][i] += _to_float(value) * weight

            other_total_this_subsector = 0.0
            for var_name, values in variables.items():
                if var_name in CORE_VARIABLES:
                    continue
                value = values[i] if i < len(values) else 0.0
                other_total_this_subsector += _to_float(value)
            other_sums[i] += other_total_this_subsector * weight

    core_means = {
        name: [
            value / weight_sums[i] if weight_sums[i] else 0.0
            for i, value in enumerate(values)
        ]
        for name, values in core_sums.items()
    }
    other_means = [
        value / weight_sums[i] if weight_sums[i] else 0.0
        for i, value in enumerate(other_sums)
    ]

    final_rows = {name: core_means[name] for name in CORE_VARIABLES}
    final_rows[OTHER_ROW_NAME] = other_means

    aggregated_df = pd.DataFrame.from_dict(final_rows, orient="index", columns=years)
    aggregated_df.loc["Total (%)"] = aggregated_df.sum(axis=0)
    return aggregated_df


def _normalize_sheet_name(raw_name, used_names):
    clean = raw_name
    for bad_char in ["\\", "/", "*", "?", ":", "[", "]"]:
        clean = clean.replace(bad_char, " ")
    clean = " ".join(clean.split())
    if not clean:
        clean = "Sheet"
    if len(clean) > 31:
        clean = clean[:31]

    candidate = clean
    idx = 2
    while candidate in used_names:
        suffix = f"_{idx}"
        max_base = 31 - len(suffix)
        candidate = f"{clean[:max_base]}{suffix}"
        idx += 1
    used_names.add(candidate)
    return candidate


def export_aggregated_country_files(output_dir=None, jrc_idees_path=None):
    """
    Create a new output folder:
    - One workbook per country
    - One sheet per sector
    - Rows per sheet: 6 attributes + Total (%)
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    if output_dir is None:
        output_dir = (
            Path(__file__).resolve().parents[1]
            / "Data"
            / "General"
            / "JRC-IDEES_final_energy_consumption_by_country_summed"
        )
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    country_codes = sorted([p.name for p in jrc_idees_path.iterdir() if p.is_dir()])
    written_files = 0
    written_sheets = 0

    for country_code in country_codes:
        blocks_by_sector = read_country_sector_blocks(country_code, jrc_idees_path=jrc_idees_path)
        if not blocks_by_sector:
            continue

        country_output = output_dir / f"{country_code}_final_energy_consumption_summed.xlsx"
        used_sheet_names = set()
        local_sheets = 0

        with pd.ExcelWriter(country_output, engine="openpyxl") as writer:
            for sector_name, blocks in blocks_by_sector.items():
                aggregated_df = aggregate_sector_blocks_to_six_rows_with_total(blocks)
                if aggregated_df is None:
                    continue

                years = list(aggregated_df.columns)
                out_df = aggregated_df.copy()
                out_df.insert(0, "Variable", out_df.index)
                out_df = out_df.reset_index(drop=True)

                header_row = [f"{country_code} final energy consumption (%)"] + years
                output_df = pd.concat(
                    [pd.DataFrame([header_row], columns=out_df.columns), out_df],
                    ignore_index=True,
                )

                excel_sheet_name = _normalize_sheet_name(sector_name, used_sheet_names)
                output_df.to_excel(writer, sheet_name=excel_sheet_name, index=False, header=False)
                local_sheets += 1

        if local_sheets > 0:
            written_files += 1
            written_sheets += local_sheets

    return output_dir, written_files, written_sheets


if __name__ == "__main__":
    out_dir, n_files, n_sheets = export_aggregated_country_files()
    print(f"Carpeta creada: {out_dir}")
    print(f"Archivos de pais escritos: {n_files}")
    print(f"Hojas de sector escritas: {n_sheets}")
