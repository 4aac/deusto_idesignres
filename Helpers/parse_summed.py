from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


CORE_VARIABLES = [
    "Lighting",
    "Air compressors",
    "Motor drives",
    "Fans and pumps",
    "Low-enthalpy heat",
]

# Update these sets/prefixes if IDEES introduces new palette variants.
COLOR_GROUPS = {
    "parent": {"4F81BD", "006EBE"},
    "detail": {"974706", "964605", "963732"},
}
COLOR_GROUP_PREFIXES = {
    "parent": ("4F81", "006E"),
    "detail": ("9747", "9646", "9637"),
}


def _default_jrc_idees_path(jrc_idees_path=None):
    if jrc_idees_path is not None:
        return Path(jrc_idees_path)
    base_dir = Path(__file__).resolve().parents[1]
    return base_dir / "Data" / "General" / "JRC-IDEES"


def read_jrc_idees_industry_files(jrc_idees_path=None, sheet_name=0):
    """
    Recorre la carpeta JRC-IDEES y lee el archivo de Industry dentro de cada subcarpeta.

    Parameters
    ----------
    jrc_idees_path : str | Path | None
        Ruta a la carpeta JRC-IDEES. Si es None, usa:
        Data/General/JRC-IDEES relativo a este archivo.
    sheet_name : str | int | None
        Hoja a leer con pandas.read_excel (por defecto, la primera: 0).

    Returns
    -------
    dict[str, pd.DataFrame]
        Diccionario {codigo_pais: dataframe}.
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    if not jrc_idees_path.exists():
        raise FileNotFoundError(f"No existe la carpeta: {jrc_idees_path}")

    industry_data = {}

    for country_dir in sorted(p for p in jrc_idees_path.iterdir() if p.is_dir()):
        industry_files = sorted(country_dir.glob("*Industry*.xlsx"))
        if not industry_files:
            continue

        # Si hubiera más de uno, usamos el primero alfabético.
        industry_file = industry_files[0]
        industry_data[country_dir.name] = pd.read_excel(industry_file, sheet_name=sheet_name)

    return industry_data


def read_jrc_idees_fec_sheets(jrc_idees_path=None):
    """
    Recorre JRC-IDEES y, en cada archivo Industry, lee el Index para cargar
    todas las hojas cuyo nombre termina en '_fec', excepto 'Ind_Summary_fec'.

    Returns
    -------
    dict[str, dict[str, pd.DataFrame]]
        Diccionario {codigo_pais: {nombre_hoja: dataframe}}.
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    if not jrc_idees_path.exists():
        raise FileNotFoundError(f"No existe la carpeta: {jrc_idees_path}")

    fec_data_by_country = {}

    for country_dir in sorted(p for p in jrc_idees_path.iterdir() if p.is_dir()):
        industry_files = sorted(country_dir.glob("*Industry*.xlsx"))
        if not industry_files:
            continue

        industry_file = industry_files[0]

        with pd.ExcelFile(industry_file) as xls:
            index_sheet_name = next(
                (name for name in xls.sheet_names if name.strip().lower() == "index"),
                None,
            )
            if index_sheet_name is None:
                raise ValueError(f"No se encontró la hoja 'index' en {industry_file}")

            index_df = pd.read_excel(xls, sheet_name=index_sheet_name, header=None)
            available_sheets = set(xls.sheet_names)

            fec_sheet_names = []
            for value in index_df.to_numpy().ravel():
                if not isinstance(value, str):
                    continue
                sheet_name = value.strip()
                if (
                    sheet_name.endswith("_fec")
                    and sheet_name != "Ind_Summary_fec"
                    and sheet_name in available_sheets
                    and sheet_name not in fec_sheet_names
                ):
                    fec_sheet_names.append(sheet_name)

            fec_data_by_country[country_dir.name] = {
                name: pd.read_excel(xls, sheet_name=name) for name in fec_sheet_names
            }

    return fec_data_by_country


def _to_clean_str(value):
    if isinstance(value, str):
        return value.strip()
    return None


def _to_float(value):
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _find_row_by_label(df, label, start_row=0, end_row=None, case_sensitive=False):
    if end_row is None:
        end_row = df.shape[0]

    target = label if case_sensitive else label.lower()
    for r in range(start_row, end_row):
        cell = df.iat[r, 0]
        if not isinstance(cell, str):
            continue
        current = cell.strip() if case_sensitive else cell.strip().lower()
        if current == target:
            return r
    return None


def _extract_year_columns(df):
    year_columns = []
    year_labels = []
    for c in range(1, df.shape[1]):
        value = df.iat[0, c]
        if pd.isna(value):
            continue
        if isinstance(value, (int, float)):
            year = int(value)
            year_columns.append(c)
            year_labels.append(year)
    return year_columns, year_labels


def _row_values(df, row_idx, col_indexes):
    values = []
    for c in col_indexes:
        value = df.iat[row_idx, c]
        if pd.isna(value):
            values.append(None)
        elif isinstance(value, (int, float)):
            values.append(float(value))
        else:
            values.append(value)
    while values and values[-1] is None:
        values.pop()
    return values


def _font_rgb_6(color):
    if color is None:
        return None
    if color.type != "rgb" or not color.rgb:
        return None
    rgb = str(color.rgb).upper()
    if len(rgb) >= 6:
        return rgb[-6:]
    return None


def _label_color(ws, row_idx):
    if ws is None:
        return None
    # DataFrame rows are 0-based, openpyxl rows are 1-based.
    return _font_rgb_6(ws.cell(row=row_idx + 1, column=1).font.color)


def _color_group(color_hex):
    if not color_hex:
        return None
    clean = str(color_hex).upper()[-6:]
    for group_name, values in COLOR_GROUPS.items():
        if clean in values:
            return group_name
    for group_name, prefixes in COLOR_GROUP_PREFIXES.items():
        if any(clean.startswith(prefix) for prefix in prefixes):
            return group_name
    return None


def _select_non_core_rows_by_color(row_items):
    """
    Color rule for non-core rows:
    - blue + next blue  -> keep blue row
    - blue + next brown -> keep following brown rows until next blue
    - any other case    -> keep current row
    """
    selected = []
    i = 0
    while i < len(row_items):
        current = row_items[i]
        current_group = _color_group(current["color"])

        if current_group == "parent":
            if i + 1 < len(row_items):
                next_group = _color_group(row_items[i + 1]["color"])
                if next_group == "parent":
                    selected.append(current)
                    i += 1
                    continue
                if next_group == "detail":
                    j = i + 1
                    while j < len(row_items) and _color_group(row_items[j]["color"]) == "detail":
                        selected.append(row_items[j])
                        j += 1
                    i = j
                    continue

            selected.append(current)
            i += 1
            continue

        selected.append(current)
        i += 1

    return selected


def _extract_block_variables(df, ws, start_row, end_row, col_indexes):
    core_rows = []
    non_core_rows = []

    for r in range(start_row, end_row + 1):
        label = _to_clean_str(df.iat[r, 0])
        if not label:
            continue

        item = {
            "label": label,
            "values": _row_values(df, r, col_indexes),
            "color": _label_color(ws, r),
        }
        if label in CORE_VARIABLES:
            core_rows.append(item)
        else:
            non_core_rows.append(item)

    variables = {}
    for item in core_rows:
        variables[item["label"]] = item["values"]

    selected_non_core = _select_non_core_rows_by_color(non_core_rows)
    for item in selected_non_core:
        variables[item["label"]] = item["values"]

    return variables


def _values_by_year(years, values):
    return {int(year): values[idx] for idx, year in enumerate(years) if idx < len(values)}


def _transpose_to_year_dict(var_to_year_values):
    year_dict = {}
    for var_name, per_year in var_to_year_values.items():
        for year, value in per_year.items():
            if year not in year_dict:
                year_dict[year] = {}
            year_dict[year][var_name] = value
    return dict(sorted(year_dict.items(), key=lambda item: item[0]))


def _iter_fec_sheet_names(xls):
    index_sheet_name = next(
        (name for name in xls.sheet_names if name.strip().lower() == "index"),
        None,
    )
    if index_sheet_name is None:
        return []

    index_df = pd.read_excel(xls, sheet_name=index_sheet_name, header=None)
    available_sheets = set(xls.sheet_names)
    fec_sheet_names = []
    for value in index_df.to_numpy().ravel():
        clean = _to_clean_str(value)
        if (
            clean
            and clean.endswith("_fec")
            and clean != "Ind_Summary_fec"
            and clean in available_sheets
            and clean not in fec_sheet_names
        ):
            fec_sheet_names.append(clean)
    return fec_sheet_names


def _find_market_rows(df):
    rows = []
    for r in range(df.shape[0]):
        label = _to_clean_str(df.iat[r, 0])
        if label and "market shares of energy uses" in label.lower():
            rows.append(r)
    return rows


def _find_next_energy_row(df, start_row):
    for r in range(start_row + 1, df.shape[0]):
        label = _to_clean_str(df.iat[r, 0])
        if label and label.lower().startswith("energy intensity"):
            return r
    return None


def _previous_non_empty_label_row(df, start_row, min_row):
    for r in range(start_row, min_row - 1, -1):
        label = _to_clean_str(df.iat[r, 0])
        if label:
            return r
    return None


def _is_excluded_sector_name(sector_name):
    if not sector_name:
        return False
    return sector_name.strip().lower() == "integrated steelworks"


def extract_market_share_blocks_by_lighting(df, ws=None):
    """
    Extrae bloques entre 'Market shares of energy uses ...' y 'Energy intensity ...'.
    Cada bloque empieza en 'Lighting' y se nombra con la fila anterior no vacía.
    Filtra atributos que contienen ' - '.
    """
    col_indexes, years = _extract_year_columns(df)
    if not col_indexes:
        return []

    blocks = []
    market_rows = _find_market_rows(df)
    for market_row in market_rows:
        energy_row = _find_next_energy_row(df, market_row)
        if energy_row is None:
            continue

        lighting_rows = []
        lighting_meta = []
        for r in range(market_row + 1, energy_row):
            label = _to_clean_str(df.iat[r, 0])
            if label and label == "Lighting":
                sector_name_row = _previous_non_empty_label_row(df, r - 1, market_row + 1)
                if sector_name_row is not None:
                    lighting_rows.append(r)
                    lighting_meta.append(
                        {
                            "lighting_row": r,
                            "sector_name_row": sector_name_row,
                            "sector_name": _to_clean_str(df.iat[sector_name_row, 0]),
                            "sector_total": _row_values(df, sector_name_row, col_indexes),
                        }
                    )

        for idx, info in enumerate(lighting_meta):
            if _is_excluded_sector_name(info["sector_name"]):
                continue

            start_row = info["lighting_row"]
            if idx + 1 < len(lighting_meta):
                # Termina justo antes del nombre del siguiente subsector
                end_row = lighting_meta[idx + 1]["sector_name_row"] - 1
            else:
                end_row = energy_row - 1

            if end_row < start_row:
                continue

            variables = _extract_block_variables(df, ws, start_row, end_row, col_indexes)

            if not variables:
                continue
            if not any(_to_float(value) > 0.0 for value in info["sector_total"]):
                continue

            variables_by_year = {name: _values_by_year(years, vals) for name, vals in variables.items()}
            blocks.append(
                {
                    "years": years,
                    "variables": variables,
                    "variables_by_year": variables_by_year,
                    "sector_name": info["sector_name"],
                    "market_row": market_row,
                    "energy_row": energy_row,
                    "lighting_row": info["lighting_row"],
                    "sector_name_row": info["sector_name_row"],
                    "sector_total": info["sector_total"],
                }
            )

    return blocks


def extract_electric_arc_market_share_block(df):
    """
    Extrae, en una hoja _fec:
    1) Variables objetivo bajo 'Electric arc' en el bloque de Market shares.
    2) Todas las variables sin indentación entre 'Electric arc' y 'Energy intensity...'.
    """
    market_row = _find_row_by_label(df, "Market shares of energy uses by subsector (%)")
    if market_row is None:
        return None

    energy_row = _find_row_by_label(
        df,
        "Energy intensity (kgoe per t of output)",
        start_row=market_row + 1,
    )
    if energy_row is None:
        return None

    electric_arc_row = _find_row_by_label(
        df,
        "Electric arc",
        start_row=market_row + 1,
        end_row=energy_row,
    )
    if electric_arc_row is not None:
        start_row = electric_arc_row + 1
        anchor_row = electric_arc_row
        anchor_label = "Electric arc"
    else:
        # Fallback para sectores sin fila "Electric arc":
        # empieza tras la primera fila de subsector no indentada encontrada.
        first_top_level_row = None
        for r in range(market_row + 1, energy_row):
            raw_label = df.iat[r, 0]
            if not isinstance(raw_label, str):
                continue
            if raw_label.strip() == "":
                continue
            if raw_label.startswith((" ", "\t")):
                continue
            first_top_level_row = r
            break
        if first_top_level_row is None:
            return None
        start_row = first_top_level_row + 1
        anchor_row = first_top_level_row
        anchor_label = str(df.iat[first_top_level_row, 0]).strip()

    target_variables = [
        "Lighting",
        "Air compressors",
        "Motor drives",
        "Fans and pumps",
        "Low-enthalpy heat",
    ]

    col_indexes, years = _extract_year_columns(df)
    if not col_indexes:
        return None

    top_level_variables = {}
    top_level_variables_by_year = {}
    for r in range(start_row, energy_row):
        raw_label = df.iat[r, 0]
        if not isinstance(raw_label, str):
            continue
        if raw_label.strip() == "":
            continue
        if raw_label.startswith((" ", "\t")):
            continue

        label = raw_label.strip()
        values = _row_values(df, r, col_indexes)
        top_level_variables[label] = values
        top_level_variables_by_year[label] = _values_by_year(years, values)

    selected_values = {}
    selected_values_by_year = {}
    for var in target_variables:
        if var in top_level_variables:
            selected_values[var] = top_level_variables[var]
            selected_values_by_year[var] = top_level_variables_by_year[var]

    selected_by_year = _transpose_to_year_dict(selected_values_by_year)
    top_level_by_year = _transpose_to_year_dict(top_level_variables_by_year)

    return {
        "years": years,
        "selected_variables": selected_values,
        "selected_variables_by_year": selected_values_by_year,
        "selected_by_year": selected_by_year,
        "top_level_variables": top_level_variables,
        "top_level_variables_by_year": top_level_variables_by_year,
        "top_level_by_year": top_level_by_year,
        "market_row": market_row,
        "electric_arc_row": electric_arc_row,
        "anchor_row": anchor_row,
        "anchor_label": anchor_label,
        "energy_intensity_row": energy_row,
    }


def extract_country_fec_electric_arc(country_code="AT", jrc_idees_path=None):
    """
    Lee todas las hojas _fec de un pais y extrae bloques definidos por
    repeticiones de 'Lighting' dentro de 'Market shares of energy uses ...'.

    Returns
    -------
    dict[str, list[dict]]
        {hoja_fec: [bloques_extraidos]}.
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    country_dir = jrc_idees_path / country_code
    if not country_dir.exists():
        raise FileNotFoundError(f"No existe la carpeta de pais: {country_dir}")

    industry_files = sorted(country_dir.glob("*Industry*.xlsx"))
    if not industry_files:
        raise FileNotFoundError(f"No se encontro archivo Industry en {country_dir}")

    industry_file = industry_files[0]
    result = {}

    with pd.ExcelFile(industry_file) as xls:
        wb = load_workbook(industry_file, data_only=True)
        fec_sheet_names = _iter_fec_sheet_names(xls)
        for sheet_name in fec_sheet_names:
            sheet_df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            blocks = extract_market_share_blocks_by_lighting(sheet_df, ws=ws)
            result[sheet_name] = blocks
        wb.close()

    return result


def print_at_example():
    extracted_by_sheet = extract_country_fec_electric_arc(country_code="AT")
    total_blocks = sum(len(blocks) for blocks in extracted_by_sheet.values())
    print(f"Hojas _fec AT detectadas: {len(extracted_by_sheet)}")
    print(f"Bloques Lighting detectados: {total_blocks}")

    for sheet_name, blocks in extracted_by_sheet.items():
        print(f"\n--- {sheet_name} --- ({len(blocks)} bloques)")
        for i, block in enumerate(blocks, start=1):
            years = block["years"]
            print(
                f"  Bloque {i}: sector='{block['sector_name']}', "
                f"filas market={block['market_row']} lighting={block['lighting_row']} "
                f"energy={block['energy_row']}, anios={years[0]}-{years[-1]}"
            )
            print(f"  Variables: {list(block['variables'].keys())}")


def export_fec_all_countries_to_excel(output_path=None, jrc_idees_path=None):
    """
    Crea un Excel con una hoja por país:
    - Fila 1: col A = '<PAIS> final energy consumption (%)', resto columnas = años.
    - Filas siguientes: variables (todas las sin indentación extraídas) y valores por año.
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    if output_path is None:
        output_path = (
            Path(__file__).resolve().parents[1]
            / "Data"
            / "General"
            / "JRC-IDEES_final_energy_consumption_all_countries.xlsx"
        )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    country_codes = sorted([p.name for p in jrc_idees_path.iterdir() if p.is_dir()])
    written_sheets = 0

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for country_code in country_codes:
            extracted_by_sheet = extract_country_fec_electric_arc(
                country_code=country_code,
                jrc_idees_path=jrc_idees_path,
            )
            if not extracted_by_sheet:
                continue

            all_years = sorted(
                {
                    year
                    for blocks in extracted_by_sheet.values()
                    for block in blocks
                    for year in block["years"]
                }
            )

            # Evita colisiones de nombres entre hojas _fec distintas
            row_data = {}
            for sheet_name, blocks in extracted_by_sheet.items():
                for block in blocks:
                    for variable, values in block["variables"].items():
                        row_name = f"{sheet_name} | {block['sector_name']} | {variable}"
                        per_year = _values_by_year(block["years"], values)
                        row_data[row_name] = {year: per_year.get(year) for year in all_years}

            if not row_data:
                continue

            country_df = pd.DataFrame.from_dict(row_data, orient="index")
            country_df = country_df.reindex(columns=all_years)

            # Primera columna = nombre de variable
            country_df.insert(0, "Variable", country_df.index)
            country_df = country_df.reset_index(drop=True)

            # Primera fila = título + años
            header_row = [f"{country_code} final energy consumption (%)"] + all_years
            output_df = pd.concat(
                [pd.DataFrame([header_row], columns=country_df.columns), country_df],
                ignore_index=True,
            )

            output_df.to_excel(writer, sheet_name=country_code, index=False, header=False)
            written_sheets += 1

    return output_path, written_sheets


def _normalize_sheet_name(raw_name, used_names):
    clean = raw_name.replace("_fec", "")
    for bad_char in ['\\', '/', '*', '?', ':', '[', ']']:
        clean = clean.replace(bad_char, " ")
    clean = " ".join(clean.split())
    if not clean:
        clean = "Sheet"
    if len(clean) > 31:
        clean = clean[:31]
    candidate = clean
    idx = 2
    while candidate in used_names:
        suffix = f"_{idx}"
        max_base = 31 - len(suffix)
        candidate = f"{clean[:max_base]}{suffix}"
        idx += 1
    used_names.add(candidate)
    return candidate


def export_fec_country_files_by_sector(output_dir=None, jrc_idees_path=None):
    """
    Crea una carpeta con un archivo Excel por país.
    En cada archivo, cada hoja es un sector (_fec) y mantiene la estructura:
    - Fila 1: '<PAIS> final energy consumption (%)' + años
    - Filas siguientes: todas las variables extraídas + valores por año
    """
    jrc_idees_path = _default_jrc_idees_path(jrc_idees_path)
    if output_dir is None:
        output_dir = (
            Path(__file__).resolve().parents[1]
            / "Data"
            / "General"
            / "JRC-IDEES_final_energy_consumption_by_country_unsummed"
        )
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    country_codes = sorted([p.name for p in jrc_idees_path.iterdir() if p.is_dir()])
    written_files = 0
    written_sheets = 0

    for country_code in country_codes:
        extracted_by_sheet = extract_country_fec_electric_arc(
            country_code=country_code,
            jrc_idees_path=jrc_idees_path,
        )
        if not extracted_by_sheet:
            continue

        country_output = output_dir / f"{country_code}_final_energy_consumption_unsummed.xlsx"
        used_sheet_names = set()
        local_sheets = 0

        with pd.ExcelWriter(country_output, engine="openpyxl") as writer:
            for sector_sheet, blocks in extracted_by_sheet.items():
                for block in blocks:
                    years = block["years"]
                    variables = block["variables"]
                    if not years or not variables:
                        continue

                    sector_df = pd.DataFrame.from_dict(variables, orient="index", columns=years)
                    sector_df.insert(0, "Variable", sector_df.index)
                    sector_df = sector_df.reset_index(drop=True)

                    # Fila de comprobacion: suma por cada año (todas las filas de variables)
                    total_values = []
                    for year in years:
                        total_values.append(pd.to_numeric(sector_df[year], errors="coerce").sum())
                    total_row_df = pd.DataFrame([["Total (%)"] + total_values], columns=sector_df.columns)
                    sector_df = pd.concat([sector_df, total_row_df], ignore_index=True)

                    header_row = [f"{country_code} final energy consumption (%)"] + years
                    output_df = pd.concat(
                        [pd.DataFrame([header_row], columns=sector_df.columns), sector_df],
                        ignore_index=True,
                    )

                    raw_sheet_name = f"{sector_sheet.replace('_fec', '')} - {block['sector_name']}"
                    excel_sheet_name = _normalize_sheet_name(raw_sheet_name, used_sheet_names)
                    output_df.to_excel(writer, sheet_name=excel_sheet_name, index=False, header=False)
                    local_sheets += 1

        if local_sheets > 0:
            written_files += 1
            written_sheets += local_sheets

    return output_dir, written_files, written_sheets


if __name__ == "__main__":
    out_dir, n_files, n_sheets = export_fec_country_files_by_sector()
    print(f"Carpeta creada: {out_dir}")
    print(f"Archivos de pais escritos: {n_files}")
    print(f"Hojas de sector escritas: {n_sheets}")


